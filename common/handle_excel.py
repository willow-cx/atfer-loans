# coding : utf-8
# @Time   :2020/6/14 15:59
# @Author :liu
# @Email  :704938465@qq.com
# @File   :handle_excel.py

from openpyxl import load_workbook


class TestData:
    pass

class HandleExcel:
    def __init__(self,filename, sheetname=None):
        self.filename = filename
        self.sheetname = sheetname

    def read_excel(self):
        test_cases = []
        header_list = []
        wb = load_workbook(self.filename)
        if self.sheetname is None:
            ws = wb.active
        else:
            ws = wb[self.sheetname]
        for row in range(1, ws.max_row+1):
            do_data = TestData()
            for column in range(1, ws.max_column+1):
                one_value = ws.cell(row, column).value
                if row == 1:
                    header_list.append(one_value)
                else:
                    setattr(do_data, str(header_list[column-1]), one_value)

            if row != 1:
                setattr(do_data, "row", row)
                test_cases.append(do_data)

        return test_cases


if __name__ == '__main__':
    do_excel = HandleExcel("D:\Pytcarm-Workspace\web_auto_learn28\data\webtestcase.xlsx")
    print(do_excel.read_excel())



